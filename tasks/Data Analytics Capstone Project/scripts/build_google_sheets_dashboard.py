"""Build a clean Google Sheets-ready dashboard workbook.

The college requires a dashboard in a tool such as Google Sheets, Tableau, or
Power BI. This workbook is designed for Google Sheets upload: the dashboard uses
stable KPI blocks and embedded chart images so it does not depend on fragile
Excel chart rendering after conversion.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TASK_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = TASK_DIR / "outputs"
ASSET_DIR = OUTPUT_DIR / "report_chart_assets"
DASHBOARD_DIR = OUTPUT_DIR / "google_sheets_dashboard"
WORKBOOK_PATH = DASHBOARD_DIR / "osmi_mental_health_google_sheets_dashboard.xlsx"

INK = "1F2933"
MUTED = "5F6B7A"
TEAL = "2F6F73"
RED = "C25746"
GOLD = "D39B5F"
VIOLET = "6D4C7D"
LIGHT = "F6F7F9"
PALE = "EEF4F4"
LINE = "D8DEE6"
WHITE = "FFFFFF"


def set_sheet_defaults(ws) -> None:
    ws.sheet_view.showGridLines = False
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, 80):
        ws.row_dimensions[row].height = 24


def style_range_border(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border


def merge_and_write(
    ws,
    cell_range: str,
    value: str,
    *,
    size: int = 11,
    bold: bool = False,
    color: str = INK,
    fill: str | None = None,
    align: str = "left",
    valign: str = "center",
) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    start, end = cell_range.split(":")
    start_col = ws[start].column
    start_row = ws[start].row
    end_col = ws[end].column
    end_row = ws[end].row
    style_range_border(ws, start_row, start_col, end_row, end_col)


def add_kpi(ws, cell_range: str, label: str, value: str, accent: str) -> None:
    start, end = cell_range.split(":")
    ws.merge_cells(cell_range)
    cell = ws[start]
    cell.value = f"{label}\n{value}"
    cell.font = Font(name="Aptos", size=13, bold=True, color=accent)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor=WHITE)
    start_col = ws[start].column
    start_row = ws[start].row
    end_col = ws[end].column
    end_row = ws[end].row
    style_range_border(ws, start_row, start_col, end_row, end_col)


def add_image(ws, image_path: Path, anchor: str, width: int) -> None:
    img = Image(str(image_path))
    scale = width / img.width
    img.width = width
    img.height = int(img.height * scale)
    ws.add_image(img, anchor)


def write_df(ws, df: pd.DataFrame, start_row: int, start_col: int) -> None:
    header_fill = PatternFill("solid", fgColor=INK)
    header_font = Font(name="Aptos", bold=True, color=WHITE)
    body_font = Font(name="Aptos", color=INK)
    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            column_name = str(df.columns[j - start_col])
            if column_name in {"company_size", "no_employees"}:
                cell.number_format = "@"
            elif column_name in {"respondents", "treatment_count"}:
                cell.number_format = "#,##0"
            elif column_name == "treatment_rate":
                cell.number_format = "0.0%"
            elif column_name == "treatment_rate_percent":
                cell.number_format = "0.0"
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    style_range_border(ws, start_row, start_col, end_row, end_col)
    for col in range(start_col, end_col + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(start_row, end_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 34)


def build_workbook() -> None:
    DASHBOARD_DIR.mkdir(parents=True, exist_ok=True)

    executive = pd.read_csv(OUTPUT_DIR / "executive_summary_metrics.csv")
    benefits = pd.read_csv(OUTPUT_DIR / "treatment_by_benefits.csv")
    care = pd.read_csv(OUTPUT_DIR / "treatment_by_care_options.csv")
    work = pd.read_csv(OUTPUT_DIR / "treatment_by_work_interfere.csv")
    company = pd.read_csv(OUTPUT_DIR / "treatment_by_company_size.csv")
    family = pd.read_csv(OUTPUT_DIR / "treatment_by_family_history.csv")
    tests = pd.read_csv(OUTPUT_DIR / "hypothesis_test_results.csv")
    model = pd.read_csv(OUTPUT_DIR / "treatment_model_performance.csv")
    coefficients = pd.read_csv(OUTPUT_DIR / "treatment_model_coefficients.csv")

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    extra_charts = wb.create_sheet("Additional_Charts")
    source = wb.create_sheet("Source_Tables")
    tests_ws = wb.create_sheet("Hypothesis_Tests")
    model_ws = wb.create_sheet("Model")
    readme = wb.create_sheet("README")

    for ws in wb.worksheets:
        set_sheet_defaults(ws)

    dashboard.freeze_panes = "A8"
    dashboard.sheet_properties.pageSetUpPr.fitToPage = True
    dashboard.page_setup.fitToWidth = 1
    dashboard.page_setup.fitToHeight = 2
    dashboard.page_margins.left = 0.25
    dashboard.page_margins.right = 0.25
    dashboard.page_margins.top = 0.35
    dashboard.page_margins.bottom = 0.35

    for col in range(1, 15):
        dashboard.column_dimensions[get_column_letter(col)].width = 12.5

    merge_and_write(
        dashboard,
        "A1:N2",
        "Workplace Mental Health Disclosure And Support Dashboard",
        size=22,
        bold=True,
        color=WHITE,
        fill=INK,
        align="center",
    )
    merge_and_write(
        dashboard,
        "A3:N4",
        (
            "OSMI Mental Health in Tech Survey | Main story: treatment-seeking is common, "
            "but workplace disclosure comfort remains much lower."
        ),
        size=12,
        color=INK,
        fill=PALE,
        align="center",
    )

    kpi_values = {row["metric"]: row["value"] for _, row in executive.iterrows()}
    kpis = [
        ("Survey responses", kpi_values["Total survey responses"], GOLD),
        ("Sought treatment", kpi_values["Sought mental health treatment"], TEAL),
        ("Know benefits", kpi_values["Know employer provides mental health benefits"], VIOLET),
        ("Any work interference", kpi_values["Report any work interference"], RED),
        ("Supervisor comfort", kpi_values["Comfortable discussing mental health with supervisor"], TEAL),
        ("Coworker comfort", kpi_values["Comfortable discussing mental health with coworkers"], GOLD),
    ]
    ranges = ["A6:C8", "D6:F8", "G6:I8", "J6:L8", "A10:C12", "D10:F12"]
    for (label, value, accent), cell_range in zip(kpis, ranges):
        add_kpi(dashboard, cell_range, label, str(value), accent)

    merge_and_write(
        dashboard,
        "G10:N12",
        (
            "Interpretation: benefit/care-option awareness, work interference, and perceived consequences "
            "are strongly associated with treatment-seeking or disclosure comfort. Results are associations, not causal claims."
        ),
        size=11,
        color=INK,
        fill=WHITE,
    )

    # Embedded chart images. These render reliably in Excel and usually import
    # cleanly into Google Sheets without text wrapping or gridline conflicts.
    add_image(dashboard, ASSET_DIR / "figure_01_treatment_overview.png", "A15", 520)
    add_image(dashboard, ASSET_DIR / "figure_02_benefits_and_care_options.png", "H15", 600)
    add_image(dashboard, ASSET_DIR / "figure_03_treatment_by_work_interference.png", "A30", 540)
    add_image(dashboard, ASSET_DIR / "figure_04_supervisor_comfort_by_consequence.png", "H30", 610)
    merge_and_write(
        dashboard,
        "A47:N49",
        (
            "Dashboard source: OSMI Mental Health in Tech Survey via Kaggle. "
            "Use Additional_Charts, Source_Tables, Hypothesis_Tests, and Model sheets for audit details."
        ),
        size=10,
        color=MUTED,
        fill=LIGHT,
    )

    extra_charts["A1"] = "Additional Dashboard Charts"
    extra_charts["A1"].font = Font(name="Aptos", size=18, bold=True, color=INK)
    extra_charts["A2"] = "Supporting visuals used in the written report and dashboard audit."
    extra_charts["A2"].font = Font(name="Aptos", size=11, color=MUTED)
    extra_charts.merge_cells("A2:N3")
    add_image(extra_charts, ASSET_DIR / "figure_05_treatment_by_company_size.png", "A5", 540)
    add_image(extra_charts, ASSET_DIR / "figure_06_model_performance.png", "H5", 540)

    source["A1"] = "Dashboard Source Tables"
    source["A1"].font = Font(name="Aptos", size=18, bold=True, color=INK)
    row = 3
    for title, df in [
        ("Treatment by benefits", benefits),
        ("Treatment by care options", care),
        ("Treatment by work interference", work),
        ("Treatment by company size", company),
        ("Treatment by family history", family),
    ]:
        source.cell(row=row, column=1, value=title)
        source.cell(row=row, column=1).font = Font(name="Aptos", size=13, bold=True, color=TEAL)
        row += 1
        write_df(source, df, row, 1)
        row += len(df) + 3

    tests_ws["A1"] = "Hypothesis Test Results"
    tests_ws["A1"].font = Font(name="Aptos", size=18, bold=True, color=INK)
    write_df(tests_ws, tests, 3, 1)

    model_ws["A1"] = "Model Results"
    model_ws["A1"].font = Font(name="Aptos", size=18, bold=True, color=INK)
    write_df(model_ws, model, 3, 1)
    model_ws["A8"] = "Top Model Coefficients"
    model_ws["A8"].font = Font(name="Aptos", size=13, bold=True, color=TEAL)
    write_df(model_ws, coefficients.head(15), 9, 1)

    readme["A1"] = "Dashboard Submission Notes"
    readme["A1"].font = Font(name="Aptos", size=18, bold=True, color=INK)
    notes = [
        "Upload this XLSX file to Google Drive and open it with Google Sheets.",
        "Submit the Google Sheets link as the required dashboard deliverable.",
        "The Dashboard sheet is intentionally image-based for stable rendering after Google Sheets conversion.",
        "Source_Tables, Hypothesis_Tests, and Model sheets provide the auditable data behind the visuals.",
        "The written report contains the same six chart snippets, and the notebook provides reproducible analysis.",
    ]
    for i, note in enumerate(notes, start=3):
        readme.cell(row=i, column=1, value=note)
        readme.cell(row=i, column=1).font = Font(name="Aptos", size=11, color=INK)
        readme.cell(row=i, column=1).alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 120

    wb.save(WORKBOOK_PATH)
    print(f"Saved Google Sheets-ready dashboard workbook: {WORKBOOK_PATH}")


if __name__ == "__main__":
    build_workbook()
