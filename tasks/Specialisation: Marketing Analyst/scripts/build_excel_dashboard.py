"""Build the Marketing Campaign Comparison Excel dashboard."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TASK_DIR = Path(__file__).resolve().parents[1]
DATA_PATH = TASK_DIR / "data" / "marketing_campaign_weekday_duration.csv"
CHART_DIR = TASK_DIR / "outputs" / "charts"
OUTPUT_PATH = (
    TASK_DIR
    / "outputs"
    / "FINAL_SUBMISSION_FILES"
    / "marketing_campaign_duration_dashboard.xlsx"
)

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
RELIABLE_CAMPAIGNS = ["(organic)", "(referral)", "<Other>", "Data Share Promo"]

INK = "24313D"
MUTED = "5F6B7A"
WHITE = "FFFFFF"
LIGHT = "F5F7FA"
LINE = "D9E0E7"
BLUE = "4666A6"
GREEN = "2F6B5F"
ORANGE = "C77C2E"
RED = "C25746"
PURPLE = "6A3D9A"


def apply_border(cell) -> None:
    cell.border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )


def apply_header(cell, fill: str = INK) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(cell)


def write_table(ws, df: pd.DataFrame, start_row: int, start_col: int) -> None:
    for col_idx, column in enumerate(df.columns, start=start_col):
        apply_header(ws.cell(start_row, col_idx, column))

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            if pd.isna(value):
                value = None
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 21

    for col_idx in range(start_col, start_col + len(df.columns)):
        width = max(
            len(str(ws.cell(row_idx, col_idx).value or ""))
            for row_idx in range(start_row, start_row + len(df) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 32)


def add_kpi(ws, start_cell: str, label: str, value: str, accent: str) -> None:
    row = ws[start_cell].row
    col = ws[start_cell].column
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 2)

    label_cell = ws.cell(row, col, label)
    label_cell.fill = PatternFill("solid", fgColor=WHITE)
    label_cell.font = Font(name="Aptos", size=10, bold=True, color=MUTED)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    value_cell = ws.cell(row + 1, col, value)
    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=accent)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(row, row + 3):
        for c in range(col, col + 3):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=WHITE)
            apply_border(ws.cell(r, c))


def load_and_summarize() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.read_csv(DATA_PATH)
    df["weekday_name"] = pd.Categorical(df["weekday_name"], categories=WEEKDAYS, ordered=True)
    df = df.sort_values(["campaign", "weekday_sort_monday_start"]).copy()

    campaign_summary = (
        df.groupby("campaign", as_index=False)
        .agg(
            total_sessions=("sessions", "sum"),
            weighted_avg_duration_minutes=(
                "avg_session_duration_minutes",
                lambda s: round(
                    (s * df.loc[s.index, "sessions"]).sum()
                    / df.loc[s.index, "sessions"].sum(),
                    2,
                ),
            ),
            median_weekday_duration_minutes=("median_session_duration_minutes", "median"),
            weighted_avg_events_per_session=(
                "avg_events_per_session",
                lambda s: round(
                    (s * df.loc[s.index, "sessions"]).sum()
                    / df.loc[s.index, "sessions"].sum(),
                    2,
                ),
            ),
            weekday_rows=("weekday_name", "count"),
        )
        .sort_values("total_sessions", ascending=False)
    )

    reliable_matrix = (
        df[df["campaign"].isin(RELIABLE_CAMPAIGNS)]
        .pivot(index="weekday_name", columns="campaign", values="avg_session_duration_minutes")
        .reindex(WEEKDAYS)
        .reset_index()
    )
    reliable_matrix.columns.name = None

    top_combinations = (
        df.sort_values("avg_session_duration_minutes", ascending=False)
        .head(12)
        .assign(
            campaign_weekday=lambda x: x["campaign"] + " - " + x["weekday_name"].astype(str),
        )[
            [
                "campaign_weekday",
                "avg_session_duration_minutes",
                "sessions",
                "median_session_duration_minutes",
            ]
        ]
    )

    return df, campaign_summary, reliable_matrix, top_combinations


def build_workbook() -> None:
    raw, campaign_summary, reliable_matrix, top_combinations = load_and_summarize()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    raw_ws = wb.create_sheet("Weekday_Data")
    summary_ws = wb.create_sheet("Campaign_Summary")
    chart_ws = wb.create_sheet("Chart_Data")
    weekday_chart_ws = wb.create_sheet("Weekday_Chart")
    sample_chart_ws = wb.create_sheet("Sample_Size_Chart")
    readme_ws = wb.create_sheet("README")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 95

    # Raw and summary sheets.
    raw_export = raw.copy()
    raw_export["weekday_name"] = raw_export["weekday_name"].astype(str)
    write_table(raw_ws, raw_export, 1, 1)
    raw_ws.freeze_panes = "A2"

    campaign_export = campaign_summary.copy()
    campaign_export.columns = [
        "Campaign",
        "Total sessions",
        "Weighted avg duration (min)",
        "Median weekday duration (min)",
        "Weighted avg events/session",
        "Weekday rows",
    ]
    write_table(summary_ws, campaign_export, 1, 1)
    summary_ws.freeze_panes = "A2"

    # Chart source tables.
    write_table(chart_ws, reliable_matrix, 1, 1)
    sample_table = campaign_export[["Campaign", "Total sessions"]].sort_values(
        "Total sessions", ascending=True
    )
    write_table(chart_ws, sample_table, 12, 1)
    top_export = top_combinations.copy()
    top_export.columns = [
        "Campaign + weekday",
        "Avg duration (min)",
        "Sessions",
        "Median duration (min)",
    ]
    write_table(chart_ws, top_export, 26, 1)
    chart_ws.freeze_panes = "A2"

    # Dashboard heading.
    dashboard["A1"] = "Marketing Campaign Comparison"
    dashboard["A1"].font = Font(name="Aptos Display", size=24, bold=True, color=INK)
    dashboard.merge_cells("A1:L1")
    dashboard["A2"] = "Weekday modeled session duration by campaign"
    dashboard["A2"].font = Font(name="Aptos", size=12, color=MUTED)
    dashboard.merge_cells("A2:L2")

    total_sessions = int(campaign_summary["total_sessions"].sum())
    referral = campaign_summary.loc[campaign_summary["campaign"] == "(referral)"].iloc[0]
    organic = campaign_summary.loc[campaign_summary["campaign"] == "(organic)"].iloc[0]
    data_share = campaign_summary.loc[
        campaign_summary["campaign"] == "Data Share Promo"
    ].iloc[0]

    add_kpi(dashboard, "A4", "Modeled sessions", f"{total_sessions:,}", BLUE)
    add_kpi(
        dashboard,
        "D4",
        "Referral avg duration",
        f"{referral['weighted_avg_duration_minutes']:.2f} min",
        BLUE,
    )
    add_kpi(
        dashboard,
        "G4",
        "Organic avg duration",
        f"{organic['weighted_avg_duration_minutes']:.2f} min",
        GREEN,
    )
    add_kpi(
        dashboard,
        "J4",
        "Data Share Promo avg",
        f"{data_share['weighted_avg_duration_minutes']:.2f} min",
        ORANGE,
    )

    dashboard["A8"] = "Main takeaways"
    dashboard["A8"].font = Font(name="Aptos Display", size=16, bold=True, color=INK)
    dashboard["A9"] = (
        "Referral sessions are consistently longer than organic sessions across the week."
    )
    dashboard["A10"] = (
        "Data Share Promo has the longest reliable weighted average, but a smaller sample."
    )
    dashboard["A11"] = (
        "Black Friday and holiday campaign averages should not be overclaimed because their samples are tiny."
    )
    for row in range(9, 12):
        dashboard[f"A{row}"].font = Font(name="Aptos", size=11, color=INK)
        dashboard[f"A{row}"].alignment = Alignment(wrap_text=True)
    dashboard.merge_cells("A9:L9")
    dashboard.merge_cells("A10:L10")
    dashboard.merge_cells("A11:L11")

    # Line chart: reliable campaigns.
    data = Reference(chart_ws, min_col=2, max_col=5, min_row=1, max_row=8)
    cats = Reference(chart_ws, min_col=1, min_row=2, max_row=8)
    line_image = ExcelImage(CHART_DIR / "reliable_campaign_weekday_duration.png")
    line_image.width = 980
    line_image.height = 560
    dashboard.add_image(line_image, "A13")

    sample_data = Reference(chart_ws, min_col=2, min_row=12, max_row=22)
    sample_cats = Reference(chart_ws, min_col=1, min_row=13, max_row=22)
    sample_image = ExcelImage(CHART_DIR / "campaign_sample_size_context.png")
    sample_image.width = 980
    sample_image.height = 580
    dashboard.add_image(sample_image, "A43")

    dashboard["A75"] = "Review note"
    dashboard["A75"].font = Font(name="Aptos Display", size=15, bold=True, color=RED)
    dashboard["A76"] = (
        "Several named campaigns have fewer than 100 modeled sessions. Treat their weekday duration as directional only."
    )
    dashboard["A76"].alignment = Alignment(wrap_text=True, vertical="top")
    dashboard["A76"].font = Font(name="Aptos", size=11, color=INK)
    dashboard.merge_cells("A76:N76")

    for col in range(1, 15):
        dashboard.column_dimensions[get_column_letter(col)].width = 14
    for row in range(1, 82):
        dashboard.row_dimensions[row].height = 22
    dashboard.row_dimensions[1].height = 34
    dashboard.row_dimensions[2].height = 24
    dashboard.sheet_properties.pageSetUpPr.fitToPage = True
    dashboard.page_setup.orientation = "landscape"
    dashboard.page_setup.fitToWidth = 1
    dashboard.page_setup.fitToHeight = 3
    dashboard.print_area = "A1:N78"

    # Dedicated large chart sheets for easy review in Excel.
    for chart_sheet in (weekday_chart_ws, sample_chart_ws):
        chart_sheet.sheet_view.showGridLines = False
        chart_sheet.sheet_view.zoomScale = 90
        chart_sheet.sheet_properties.pageSetUpPr.fitToPage = True
        chart_sheet.page_setup.orientation = "landscape"
        chart_sheet.page_setup.fitToWidth = 1
        chart_sheet.page_setup.fitToHeight = 1
        for col in range(1, 15):
            chart_sheet.column_dimensions[get_column_letter(col)].width = 14
        for row in range(1, 34):
            chart_sheet.row_dimensions[row].height = 24

    weekday_chart_ws["A1"] = "Weekday Average Session Duration"
    weekday_chart_ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=INK)
    weekday_chart_ws.merge_cells("A1:N1")
    weekday_chart_ws["A2"] = "X-axis: weekday. Y-axis: average modeled session duration in minutes."
    weekday_chart_ws["A2"].font = Font(name="Aptos", size=12, color=MUTED)
    weekday_chart_ws.merge_cells("A2:N2")
    large_line = ExcelImage(CHART_DIR / "reliable_campaign_weekday_duration.png")
    large_line.width = 1040
    large_line.height = 590
    weekday_chart_ws.add_image(large_line, "A4")
    weekday_chart_ws.print_area = "A1:N33"

    sample_chart_ws["A1"] = "Campaign Sample Size Context"
    sample_chart_ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=INK)
    sample_chart_ws.merge_cells("A1:N1")
    sample_chart_ws["A2"] = "X-axis: total modeled sessions on a log scale. Y-axis: campaign."
    sample_chart_ws["A2"].font = Font(name="Aptos", size=12, color=MUTED)
    sample_chart_ws.merge_cells("A2:N2")
    large_sample = ExcelImage(CHART_DIR / "campaign_sample_size_context.png")
    large_sample.width = 1040
    large_sample.height = 600
    sample_chart_ws.add_image(large_sample, "A4")
    sample_chart_ws.print_area = "A1:N33"

    # README sheet.
    readme_lines = [
        ("A1", "Marketing Campaign Comparison Dashboard", 18, True),
        ("A3", "Source", 13, True),
        ("A4", "BigQuery export from tc-da-1.turing_data_analytics.raw_events.", 11, False),
        ("A6", "Metric definition", 13, True),
        (
            "A7",
            "A modeled session starts after a new day or a gap longer than 30 minutes. Duration is last event time minus first event time.",
            11,
            False,
        ),
        ("A9", "Important caution", 13, True),
        (
            "A10",
            "Campaigns with very small session counts should not be used for strong performance claims.",
            11,
            False,
        ),
        ("A12", "Workbook tabs", 13, True),
        (
            "A13",
            "Dashboard: main charts and findings. Weekday_Chart and Sample_Size_Chart: large chart views. Weekday_Data: BigQuery result. Campaign_Summary: weighted campaign-level summary. Chart_Data: helper ranges for charts.",
            11,
            False,
        ),
    ]
    for cell, value, size, bold in readme_lines:
        readme_ws[cell] = value
        readme_ws[cell].font = Font(name="Aptos", size=size, bold=bold, color=INK)
        readme_ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    readme_ws.column_dimensions["A"].width = 110

    wb.save(OUTPUT_PATH)

    # Verify the workbook can be opened again.
    load_workbook(OUTPUT_PATH, read_only=True).close()
    print(f"Workbook written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_workbook()
