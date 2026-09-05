"""Build a Google Sheets-ready Excel workbook with native charts."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TASK_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = TASK_DIR / "outputs"
WORKBOOK_PATH = OUTPUT_DIR / "time_to_purchase_dashboard.xlsx"

INK = "24313D"
MUTED = "5F6B7A"
BLUE = "2878B5"
ORANGE = "E76F51"
GREEN = "3A9D5D"
PURPLE = "8561A8"
LIGHT = "F5F7FA"
PALE_BLUE = "EAF3FA"
WHITE = "FFFFFF"
LINE = "D9E0E7"


def apply_header(cell, fill: str = INK) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_border(cell) -> None:
    cell.border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )


def write_table(ws, df: pd.DataFrame, start_row: int, start_col: int) -> None:
    for j, column in enumerate(df.columns, start=start_col):
        cell = ws.cell(start_row, j, column)
        apply_header(cell)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            if pd.isna(value):
                value = None
            cell = ws.cell(i, j, value)
            cell.font = Font(name="Aptos", size=11, color=INK)
            cell.alignment = Alignment(vertical="center")
            apply_border(cell)
        ws.row_dimensions[i].height = 21
    for col in range(start_col, start_col + len(df.columns)):
        max_len = max(
            len(str(ws.cell(row, col).value or ""))
            for row in range(start_row, start_row + len(df) + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 28)


def add_kpi(ws, cell_range: str, label: str, value: object, number_format: str, accent: str) -> None:
    ws.merge_cells(cell_range)
    start = cell_range.split(":")[0]
    cell = ws[start]
    cell.value = f"{label}\n"
    cell.font = Font(name="Aptos", size=11, bold=True, color=MUTED)
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    apply_border(cell)

    start_col = ws[start].column
    start_row = ws[start].row
    end_col = ws[cell_range.split(":")[1]].column
    end_row = ws[cell_range.split(":")[1]].row
    value_cell = ws.cell(start_row + 1, start_col)
    ws.unmerge_cells(cell_range)
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
    label_cell = ws.cell(start_row, start_col)
    label_cell.value = label
    label_cell.font = Font(name="Aptos", size=11, bold=True, color=MUTED)
    label_cell.fill = PatternFill("solid", fgColor=WHITE)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell = ws.cell(start_row + 1, start_col)
    value_cell.value = value
    value_cell.number_format = number_format
    value_cell.font = Font(name="Aptos Display", size=22, bold=True, color=accent)
    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            apply_border(ws.cell(row, col))


def style_chart(chart, title: str, y_title: str | None, width: float, height: float) -> None:
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = None
    chart.width = width
    chart.height = height
    chart.legend.position = "b"
    chart.style = 2
    chart.y_axis.majorGridlines = None
    chart.y_axis.numFmt = "0"
    chart.x_axis.txPr = None
    chart.y_axis.txPr = None


def main() -> None:
    daily = pd.read_csv(TASK_DIR / "data" / "daily_time_to_purchase.csv", parse_dates=["event_date"])
    user = pd.read_csv(TASK_DIR / "data" / "user_level_time_to_purchase.csv", parse_dates=["event_date"])

    device = (
        user.groupby("device_category")
        .agg(
            purchasing_users=("user_pseudo_id", "count"),
            average_minutes=("duration_to_purchase_minutes", "mean"),
            median_minutes=("duration_to_purchase_minutes", "median"),
            p90_minutes=("duration_to_purchase_minutes", lambda s: s.quantile(0.90)),
        )
        .round(2)
        .reset_index()
    )
    country = (
        user.groupby("country")
        .agg(
            purchasing_users=("user_pseudo_id", "count"),
            average_minutes=("duration_to_purchase_minutes", "mean"),
            median_minutes=("duration_to_purchase_minutes", "median"),
            p90_minutes=("duration_to_purchase_minutes", lambda s: s.quantile(0.90)),
        )
        .round(2)
        .sort_values("purchasing_users", ascending=False)
        .head(15)
        .reset_index()
    )
    summary_columns = {
        "device_category": "Device",
        "country": "Country",
        "purchasing_users": "Purchasers",
        "average_minutes": "Average (min)",
        "median_minutes": "Median (min)",
        "p90_minutes": "P90 (min)",
    }
    device = device.rename(columns=summary_columns)
    country = country.rename(columns=summary_columns)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    daily_ws = wb.create_sheet("Daily_Data")
    segment_ws = wb.create_sheet("Segment_Summary")
    user_ws = wb.create_sheet("User_Level_Data")
    readme = wb.create_sheet("README")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.view = "normal"
        ws.sheet_view.zoomScale = 95

    # Source sheets.
    daily_export = daily.copy()
    daily_export["event_date"] = daily_export["event_date"].dt.date
    daily_export.columns = [
        "Date",
        "Purchasing users",
        "Daily average minutes",
        "Daily median minutes",
        "P25 minutes",
        "P75 minutes",
        "P90 minutes",
        "Minimum minutes",
        "Maximum minutes",
    ]
    write_table(daily_ws, daily_export, 1, 1)
    daily_ws.freeze_panes = "A2"
    for cell in daily_ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"

    user_export = user.copy()
    user_export["event_date"] = user_export["event_date"].dt.date
    user_export.columns = [
        "Date",
        "User ID",
        "Country",
        "Device category",
        "First arrival time",
        "First purchase time",
        "Duration minutes",
    ]
    write_table(user_ws, user_export, 1, 1)
    user_ws.freeze_panes = "A2"
    for cell in user_ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    user_ws.column_dimensions["B"].width = 18
    user_ws.column_dimensions["E"].width = 30
    user_ws.column_dimensions["F"].width = 30

    segment_ws["A1"] = "Segment Summary"
    segment_ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=INK)
    segment_ws.merge_cells("A1:E1")
    segment_ws["A2"] = "Device"
    segment_ws["A2"].font = Font(name="Aptos Display", size=16, bold=True, color=BLUE)
    write_table(segment_ws, device, 3, 1)
    country_start = 23
    segment_ws.cell(country_start, 1, "Top Countries by Purchasing Users")
    segment_ws.cell(country_start, 1).font = Font(name="Aptos Display", size=16, bold=True, color=PURPLE)
    write_table(segment_ws, country, country_start + 2, 1)
    segment_ws.freeze_panes = None
    segment_ws.sheet_view.zoomScale = 100
    segment_ws.sheet_properties.pageSetUpPr.fitToPage = True
    segment_ws.page_setup.orientation = "landscape"
    segment_ws.page_setup.fitToWidth = 1
    segment_ws.page_setup.fitToHeight = 2
    segment_ws.print_area = "A1:S42"
    segment_ws.column_dimensions["A"].width = 24
    for col in ("B", "C", "D", "E"):
        segment_ws.column_dimensions[col].width = 20
    for col in range(7, 20):
        segment_ws.column_dimensions[get_column_letter(col)].width = 10

    # Dedicated chart ranges avoid merged cells and formatted table headers.
    segment_ws["U2"] = "Device"
    segment_ws["V2"] = "Median minutes"
    for row, values in enumerate(device[["Device", "Median (min)"]].itertuples(index=False), start=3):
        segment_ws.cell(row, 21, values[0])
        segment_ws.cell(row, 22, values[1])

    segment_ws["U8"] = "Country"
    segment_ws["V8"] = "Median minutes"
    short_country_names = {
        "United States": "US",
        "United Kingdom": "UK",
    }
    for row, values in enumerate(
        country[["Country", "Median (min)"]].head(10).itertuples(index=False),
        start=9,
    ):
        segment_ws.cell(row, 21, short_country_names.get(values[0], values[0]))
        segment_ws.cell(row, 22, values[1])

    segment_ws.column_dimensions["U"].width = 18
    segment_ws.column_dimensions["V"].width = 16

    # Dashboard structure.
    dashboard.sheet_view.showGridLines = False
    dashboard.freeze_panes = None
    dashboard.sheet_view.zoomScale = 100
    dashboard.sheet_properties.pageSetUpPr.fitToPage = True
    dashboard.page_setup.orientation = "landscape"
    dashboard.page_setup.fitToWidth = 1
    dashboard.page_setup.fitToHeight = 2
    dashboard.page_margins.left = 0.25
    dashboard.page_margins.right = 0.25
    dashboard.page_margins.top = 0.35
    dashboard.page_margins.bottom = 0.35
    dashboard.print_area = "A1:R60"
    for col in range(1, 19):
        dashboard.column_dimensions[get_column_letter(col)].width = 10
    for row in range(1, 61):
        dashboard.row_dimensions[row].height = 24

    dashboard.merge_cells("A1:R2")
    title = dashboard["A1"]
    title.value = "Time to Purchase Dashboard"
    title.font = Font(name="Aptos Display", size=24, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=INK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    dashboard.merge_cells("A3:R4")
    subtitle = dashboard["A3"]
    subtitle.value = (
        "Same-day purchaser journey | First recorded event to first same-day purchase | "
        "2020-11-01 to 2021-01-31"
    )
    subtitle.font = Font(name="Aptos", size=12, color=INK)
    subtitle.fill = PatternFill("solid", fgColor=PALE_BLUE)
    subtitle.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    user_end = len(user) + 1
    add_kpi(dashboard, "A6:D9", "Purchasing journeys", f"=COUNTA(User_Level_Data!G2:G{user_end})", "#,##0", BLUE)
    add_kpi(dashboard, "F6:I9", "Overall median", f"=MEDIAN(User_Level_Data!G2:G{user_end})", "0.00 \"min\"", GREEN)
    add_kpi(dashboard, "K6:N9", "Overall average", f"=AVERAGE(User_Level_Data!G2:G{user_end})", "0.00 \"min\"", ORANGE)
    add_kpi(
        dashboard,
        "O6:R9",
        "90th percentile",
        round(user["duration_to_purchase_minutes"].quantile(0.90), 2),
        "0.00 \"min\"",
        PURPLE,
    )

    dashboard.merge_cells("A11:R12")
    insight = dashboard["A11"]
    insight.value = (
        "Decision: track median time to purchase as the headline KPI. "
        "Use purchaser count and p90 to monitor volume and the long tail."
    )
    insight.font = Font(name="Aptos", size=13, bold=True, color=INK)
    insight.fill = PatternFill("solid", fgColor=LIGHT)
    insight.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Daily median chart with IQR.
    daily_rows = len(daily_export) + 1
    trend = LineChart()
    trend.add_data(Reference(daily_ws, min_col=4, max_col=6, min_row=1, max_row=daily_rows), titles_from_data=True)
    trend.set_categories(Reference(daily_ws, min_col=1, min_row=2, max_row=daily_rows))
    style_chart(trend, "Daily Median and Middle 50% (min)", None, 15.0, 8.2)
    trend.series[0].graphicalProperties.line.solidFill = BLUE
    trend.series[0].graphicalProperties.line.width = 18000
    trend.series[1].graphicalProperties.line.solidFill = "A8CBE2"
    trend.series[1].graphicalProperties.line.width = 10000
    trend.series[2].graphicalProperties.line.solidFill = "A8CBE2"
    trend.series[2].graphicalProperties.line.width = 10000
    dashboard.add_chart(trend, "A14")

    avg_chart = LineChart()
    avg_chart.add_data(Reference(daily_ws, min_col=3, max_col=4, min_row=1, max_row=daily_rows), titles_from_data=True)
    avg_chart.set_categories(Reference(daily_ws, min_col=1, min_row=2, max_row=daily_rows))
    style_chart(avg_chart, "Daily Average vs Median (min)", None, 15.0, 8.2)
    avg_chart.series[0].graphicalProperties.line.solidFill = ORANGE
    avg_chart.series[0].graphicalProperties.line.width = 18000
    avg_chart.series[1].graphicalProperties.line.solidFill = BLUE
    avg_chart.series[1].graphicalProperties.line.width = 18000
    dashboard.add_chart(avg_chart, "J14")

    volume = BarChart()
    volume.type = "col"
    volume.add_data(Reference(daily_ws, min_col=2, min_row=1, max_row=daily_rows), titles_from_data=True)
    volume.set_categories(Reference(daily_ws, min_col=1, min_row=2, max_row=daily_rows))
    style_chart(volume, "Daily Same-Day Purchasers", None, 15.0, 8.2)
    volume.series[0].graphicalProperties.solidFill = GREEN
    volume.gapWidth = 65
    volume.legend = None
    dashboard.add_chart(volume, "A36")

    device_chart = BarChart()
    device_chart.type = "col"
    device_chart.add_data(Reference(segment_ws, min_col=4, min_row=3, max_row=3 + len(device)), titles_from_data=True)
    device_chart.set_categories(Reference(segment_ws, min_col=1, min_row=4, max_row=3 + len(device)))
    style_chart(device_chart, "Median Duration by Device (min)", None, 15.0, 8.2)
    device_chart.series[0].graphicalProperties.solidFill = BLUE
    device_chart.gapWidth = 85
    device_chart.dLbls = DataLabelList()
    device_chart.dLbls.showVal = True
    device_chart.dLbls.showCatName = False
    device_chart.dLbls.showSerName = False
    device_chart.dLbls.showLegendKey = False
    device_chart.dLbls.dLblPos = "outEnd"
    device_chart.dLbls.numFmt = "0.00"
    device_chart.legend = None
    dashboard.add_chart(device_chart, "J36")

    country_chart = BarChart()
    country_chart.type = "col"
    country_chart.add_data(
        Reference(segment_ws, min_col=22, min_row=8, max_row=18),
        titles_from_data=True,
    )
    country_chart.set_categories(
        Reference(segment_ws, min_col=21, min_row=9, max_row=18)
    )
    style_chart(country_chart, "Median Duration: Top 10 Countries (min)", None, 20.0, 10.0)
    country_chart.x_axis.title = "Country"
    country_chart.y_axis.title = "Median duration (minutes)"
    country_chart.series[0].graphicalProperties.solidFill = PURPLE
    country_chart.gapWidth = 65
    country_chart.dLbls = DataLabelList()
    country_chart.dLbls.showVal = True
    country_chart.dLbls.showCatName = False
    country_chart.dLbls.showSerName = False
    country_chart.dLbls.showLegendKey = False
    country_chart.dLbls.dLblPos = "outEnd"
    country_chart.dLbls.numFmt = "0.00"
    country_chart.y_axis.scaling.min = 0
    country_chart.y_axis.scaling.max = 30
    country_chart.legend = None
    segment_ws.add_chart(country_chart, "G25")

    device_segment_chart = BarChart()
    device_segment_chart.type = "col"
    device_segment_chart.add_data(
        Reference(segment_ws, min_col=22, min_row=2, max_row=5),
        titles_from_data=True,
    )
    device_segment_chart.set_categories(
        Reference(segment_ws, min_col=21, min_row=3, max_row=5)
    )
    style_chart(device_segment_chart, "Median Duration by Device (min)", None, 20.0, 8.5)
    device_segment_chart.x_axis.title = "Device"
    device_segment_chart.y_axis.title = "Median duration (minutes)"
    device_segment_chart.series[0].graphicalProperties.solidFill = BLUE
    device_segment_chart.gapWidth = 85
    device_segment_chart.dLbls = DataLabelList()
    device_segment_chart.dLbls.showVal = True
    device_segment_chart.dLbls.showCatName = False
    device_segment_chart.dLbls.showSerName = False
    device_segment_chart.dLbls.showLegendKey = False
    device_segment_chart.dLbls.dLblPos = "outEnd"
    device_segment_chart.dLbls.numFmt = "0.00"
    device_segment_chart.y_axis.scaling.min = 0
    device_segment_chart.y_axis.scaling.max = 25
    device_segment_chart.legend = None
    segment_ws.add_chart(device_segment_chart, "G3")

    # README sheet.
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 105
    readme["A1"] = "Workbook Guide"
    readme["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=INK)
    guide = [
        ("Purpose", "Google Sheets-ready visualization workbook for the Product Analyst time-to-purchase project."),
        ("Arrival definition", "Because the task asks for first arrival on a given day, the first recorded event for each user-date is used as the arrival timestamp."),
        ("Date/time limitation", "Dates are derived from event timestamps in UTC. Results may change under a business-local timezone or session-based definition."),
        ("Overall average", "74.68 minutes across 4,794 user-level journeys."),
        ("Overall median", "19.07 minutes across 4,794 user-level journeys."),
        ("Daily statistic", "67.66 minutes is the average of the 92 daily average durations, not the overall user-level average."),
        ("Recommended KPI", "Weekly median time to purchase, supported by purchaser count, p75, and p90."),
        ("Google Sheets", "Upload this XLSX to Google Drive and open it with Google Sheets. Check that native charts imported correctly."),
    ]
    for row, (label, value) in enumerate(guide, start=3):
        readme.cell(row, 1, label).font = Font(name="Aptos", size=11, bold=True, color=INK)
        readme.cell(row, 2, value).font = Font(name="Aptos", size=11, color=INK)
        readme.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        readme.row_dimensions[row].height = 36

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_PATH)

    # Structural verification.
    check = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    assert check.sheetnames == ["Dashboard", "Daily_Data", "Segment_Summary", "User_Level_Data", "README"]
    assert len(check["Daily_Data"]["A"]) == len(daily) + 1
    assert len(check["User_Level_Data"]["A"]) == len(user) + 1
    assert len(check["Dashboard"]._charts) == 4
    assert len(check["Segment_Summary"]._charts) == 2
    print(f"Saved workbook: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
